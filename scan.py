
# import os
# import PyPDF2
# import openpyxl  # Library for creating and writing to Excel files

# def search_pdfs_in_folder(rank, category):
  
#   folder_path = "files".encode('utf-8').decode('utf-8')  # Path to the folder containing PDF files
#   excel_file_name = "results.xlsx"  # Name of the output Excel file
#   wb = openpyxl.Workbook()  # Create a new workbook
#   sheet = wb.active  # Get the active sheet

#   # Add a header row to the Excel sheet (optional)
#   sheet.cell(row=1, column=1).value = "Filename"

#   for filename in os.listdir(folder_path):
#     if filename.endswith(".pdf"):
#       full_path = os.path.join(folder_path, filename)
#       found = search_pdf_for_word(full_path, category, rank)
#       if found:
#         sheet.append([filename])  # Append filename to a new row

#   wb.save(excel_file_name)  # Save the workbook as the specified Excel file

# def search_pdf_for_word(pdf_path, search_word, rank):

#   with open(pdf_path, 'rb') as pdf_file:
#     pdf_reader = PyPDF2.PdfReader(pdf_file)
#     for page_num in range(len(pdf_reader.pages)):
#       page = pdf_reader.pages[page_num]
#       page_text = page.extract_text()
#       lines = page_text.splitlines()  # Split into lines

#       for line in lines:
#         if search_word.lower() in line.lower():  # Check for "SC" (case-insensitive)
#           parts = line.strip().split()

#           if len(parts) >= 5:
#             try:
#               closing_rank = int(parts[1])  # Assuming closing rank is the second element
#               if closing_rank > rank:  # Check if closing rank is greater than 10000
#                 return True  # Condition met, exit function
#             except ValueError:
#               pass  # Skip lines with non-numeric closing rank

#   return False  # Condition not met in this file



import os
import PyPDF2
import openpyxl  # Library for creating and writing to Excel files
import re

# def search_pdfs_in_folder(rank, category):
#     folder_path = "files".encode('utf-8').decode('utf-8')  # Path to the folder containing PDF files
#     excel_file_name = "results.xlsx"  # Name of the output Excel file
#     wb = openpyxl.Workbook()  # Create a new workbook
#     sheet = wb.active  # Get the active sheet

#     # Add a header row to the Excel sheet (optional)
#     sheet.cell(row=1, column=1).value = "Filename"

#     for filename in os.listdir(folder_path):
#         if filename.endswith(".pdf"):
#             full_path = os.path.join(folder_path, filename)
           
#             found = search_pdf_for_word(full_path, category, rank)
#             if found:
#                 print(filename)
#                 sheet.append([filename])  # Append filename to a new row

#     wb.save(excel_file_name)  # Save the workbook as the specified Excel file
#     return excel_file_name  # Return the name of the created Excel file

# def search_pdf_for_word(pdf_path, search_word, rank):
#     with open(pdf_path, 'rb') as pdf_file:
#         pdf_reader = PyPDF2.PdfReader(pdf_file)
#         for page_num in range(len(pdf_reader.pages)):
#             page = pdf_reader.pages[page_num]
#             page_text = page.extract_text()
#             lines = page_text.splitlines()  # Split into lines

#             for line in lines:
#                 if search_word.lower() in line.lower():  # Check for the category (case-insensitive)
#                     parts = line.strip().split()

#                     if len(parts) >= 5:
#                         try:
#                             closing_rank = int(parts[1])  # Assuming closing rank is the second element
#                             if closing_rank > rank:  # Check if closing rank is greater than the provided rank
#                                 return True  # Condition met, exit function
#                         except ValueError:
#                             pass  # Skip lines with non-numeric closing rank

#     return False  # Condition not met in this file

import os
import openpyxl
import PyPDF2
import re

def search_pdfs_in_folder(rank, category):
    folder_path = "files".encode('utf-8').decode('utf-8')  # Path to the folder containing PDF files
    excel_file_name = "results.xlsx"  # Name of the output Excel file
    wb = openpyxl.Workbook()  # Create a new workbook
    sheet = wb.active  # Get the active sheet

    # Add a header row to the Excel sheet (optional)
    sheet.cell(row=1, column=1).value = "Filename"

    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            full_path = os.path.join(folder_path, filename)
           
            found = search_pdf_for_word(full_path, category, rank)
            if found:
                print(filename)
                sheet.append([filename])  # Append filename to a new row

    wb.save(excel_file_name)  # Save the workbook as the specified Excel file
    return excel_file_name  # Return the name of the created Excel file

def search_pdf_for_word(pdf_path, search_word, rank):
    word_pattern = r'\b' + re.escape(search_word) + r'\b'  # Create a regex pattern for exact word match
    
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            lines = page_text.splitlines()  # Split into lines

            for line in lines:
                match = re.search(word_pattern, line, re.IGNORECASE)
                if match:
                    start_index = match.end()
                    following_text = line[start_index:start_index+4]
                    
                    if "pwd" not in following_text.lower():  # Ensure "pwd" is not in the next 4 characters
                        parts = line.strip().split()

                        if len(parts) >= 5:
                            try:
                                closing_rank = int(parts[1])  # Assuming closing rank is the second element
                                if closing_rank > rank:  # Check if closing rank is greater than the provided rank
                                    return True  # Condition met, exit function
                            except ValueError:
                                pass  # Skip lines with non-numeric closing rank

    return False  # Condition not met in this file




  # Search for closing ranks greater than 10000 for SC category